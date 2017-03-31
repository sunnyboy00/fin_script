# -*- coding:gbk -*-
import sys
import re
import os
import time
import datetime
import tushare as ts
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
from internal.common import *

#��ȡ����еļ۸�
def get_xg_fx():
	sheet_st = "Sheet"
	wkfile = "..\\Data\\entry\\xingu\\faxing.xlsx"
	wkfile1 = "..\\Data\\entry\\xingu\\faxing_re.xlsx"

	wb = load_workbook(wkfile)
	ws = wb.get_sheet_by_name(sheet_st)	
	first_list = []
	item_list = []
	xg_df = pd.DataFrame()
	for rx in range(1,ws.max_row+1):
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value

		if rx==1:
			first_list = [w1,w2,w3,w4,w5,w6]
			continue

		temp_list = [w1,w2,w3,w4,w5,w6]
		item_list.append(temp_list)
	df1 = pd.DataFrame(item_list, columns=first_list)
	xg_df = xg_df.append(df1)

	item_list = []
	wb = load_workbook(wkfile1)
	ws = wb.get_sheet_by_name(sheet_st)	
	for rx in range(2,ws.max_row+1):
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value

		temp_list = [w1,w2,w3,w4,w5,w6]
		item_list.append(temp_list)
	if len(item_list)>0:
		df1 = pd.DataFrame(item_list, columns=first_list)
		xg_df = xg_df.append(df1)

	xg_df = xg_df.set_index('code')
	#print xg_df.ix['001979'][3]

	return xg_df

# Main
cmp_string = "20150201"
base_date = datetime.datetime.strptime(cmp_string, '%Y%m%d').date()

prepath = "..\\Data\\"
prepath1 = "..\\Data\\entry\\cixin\\"
LOOP_COUNT = 0
df = None

xg_df = get_xg_fx()
xg_list = list(xg_df.index)

#�õ�basic info
while LOOP_COUNT<3:
	try:
		df = ts.get_stock_basics()
	except:
		LOOP_COUNT += 1
		time.sleep(0.5)
	else:
		break;
if df is None:
	print "Timeout to get stock basic info"
	exit(0)
df1 = df.sort_values(['timeToMarket'], 0, False)
#df1 = df1[332:333]

index = -1
wb = Workbook()
# grab the active worksheet
ws = wb.active
strline = u'����,����,���м�,����,�����,����,������,������,����ZT,����CZT,��ͨ�ɱ�,��ͨ��ֵ,�ܹɱ�,����ֵ,�ⵥ����,�ⵥ��ͨ��,������,����DT'
strObj = strline.split(u',')
ws.append(strObj)
#�����������иı�
ws.auto_filter.ref = "A1:R1"
excel_row = 2
for code,row in df1.iterrows():
	stockInfo = []
	index += 1
	name = row[0].decode('utf8')
	ipo_date = row['timeToMarket']
	liutong_gb = row['outstanding']
	zong_gb = row['totals']
	#print type(ipo_date) ��Ȼ�� long ����
	trade_string = str(long(ipo_date))
	trade_date = datetime.datetime.strptime(trade_string, '%Y%m%d').date()
	delta = trade_date - base_date
	#print (index+1),code,delta.days,trade_date,base_date
	#print (index+1),code,name,trade_date	
	if delta.days<0:
		break

	#���ÿֻ����ÿ�콻������
	LOOP_COUNT = 0
	tddf = None
	while LOOP_COUNT<3:
		try:
			tddf = ts.get_k_data(code, autype='bfq')
			#tddf = ts.get_k_data(code)
		except:
			LOOP_COUNT += 1
			time.sleep(0.5)
		else:
			break;
	if tddf is None:
		print "Timeout to get k data of " + code +", Quit"
		exit(0)

	b_open = 0
	b_break = 0
	b_log_cxkb = 0
	yzzt_day = 0
	last_close = 0.0
	td_total = len(tddf)
	fengban_vol = 0		#�������
	fengliu_prop = 0.0
	last_day_vol = 0.0
	turnover = 0.0
	cxkb_date = 0
	kbzt_days = 0		#�����������ͣ�����Ƿ�����ͣ
	kbczt_days = 0
	kbdt_days = 0		#�����������ͣ�����Ƿ�����ͣ
	for tdidx,tdrow in tddf.iterrows():
		open = tdrow[1]
		close = tdrow[2]
		high = tdrow['high']
		low = tdrow['low']
		last_day_vol = tdrow['volume']
		#high == low ��ζYZZT
		#print code,name,kbczt_days,kbzt_days,tdrow['date'],high,low
		if high==low:
			yzzt_day += 1
			last_close = close
			if b_open==1:
				kbczt_days += 1
				kbzt_days += 1
			continue

		#�¹ɵ�һ����� high!=low
		if yzzt_day!=0:
			b_open = 1
			opn_date_str = tdrow['date']
			#������ͣ����
			zt_price = last_close * 1.0992
			dt_price = last_close * 0.9005
			#print code,name,last_close,close,zt_price,high,kbczt_days,kbzt_days
			if (close>=zt_price or high>=zt_price) and (kbdt_days==0):
				#ͨ�����㿪������ͣ����ͣ�����������ж�
				#�����һ�����ͣ���ڶ�����ͣ������Ϊ������ͣ
				if kbczt_days==kbzt_days and close>=zt_price:
					kbzt_days += 1
				kbczt_days += 1
			elif close<=dt_price and kbczt_days==0 and kbzt_days==0:
				kbdt_days += 1
			else:
				b_break = 1
			b_log_cxkb = 1
		else:
			#��������¹ɣ������߿ڡ����Ϲɷݵ�
			if open<=close and close==high and low==open:
				yzzt_day += 1
			else:
				b_open = 1
				b_break = 1
				b_log_cxkb = 1
				opn_date_str = tdrow['date']
		last_close = close
		
		#���ܿ��壬���ǿ��ܻ����������ZT or DT��������Ҫ��¼��������
		if b_log_cxkb==1:
			if cxkb_date==0:
				opn_date_int = ''.join(opn_date_str.split('-'))
				cxkb_date = int(opn_date_int)
		if b_break==1:
			break

	#ֻ��û�д򿪵�CX���ż����⼸�����ݣ��򿪵ľͺ���
	if b_open==0:
		LOOP_COUNT = 0
		trdf = None
		while LOOP_COUNT<3:
			try:
				trdf = ts.get_realtime_quotes(code)
			except:
				LOOP_COUNT += 1
				time.sleep(0.5)
			else:
				break;
		if trdf is None:
			print "Timeout to get real time quotes"
			exit(0)

		volstr = trdf.iloc[0,10]
		if volstr.isdigit() is True:
			fengban_vol = int(volstr)
			fengliu_prop = fengban_vol/(liutong_gb*10000)
			turnover = last_day_vol/(liutong_gb*10000)

	#׷������,��ͨ��ֵ������ֵ
	liutong_sz = liutong_gb*last_close
	zong_sz = zong_gb*last_close
	stockInfo.append(code)
	stockInfo.append(name)
	if xg_df is None:
		stockInfo.append(None)
	else:
		if code in xg_list:
			stockInfo.append(xg_df.ix[code][3])
		else:
			print code,name, "No in list"
	stockInfo.append(b_open)
	stockInfo.append(yzzt_day)
	stockInfo.append(last_close)
	stockInfo.append(ipo_date)
	stockInfo.append(cxkb_date)
	stockInfo.append(kbzt_days)
	stockInfo.append(kbczt_days)
	stockInfo.append(liutong_gb)
	stockInfo.append(round(liutong_sz,2))
	stockInfo.append(zong_gb)
	stockInfo.append(round(zong_sz,2))
	stockInfo.append(fengban_vol)
	stockInfo.append(round(fengliu_prop,2))
	stockInfo.append(round(turnover,2))
	stockInfo.append(kbdt_days)
	#print stockInfo

	k = 0
	ascid = 65
	number = len(stockInfo)
	for k in range(0,number):
		cell = chr(ascid+k) + str(excel_row)
		ws[cell] = stockInfo[k]
	excel_row += 1

filexlsx = prepath + "cixin_analyze.xlsx"
wb.save(filexlsx)

today = datetime.date.today()
cur=datetime.datetime.now()
qdate = '%04d-%02d-%02d' %(today.year, today.month, today.day)
filexlsx1 = prepath1 + "cx_anly_"+ qdate
filexlsx1 = '%s#%02d-%02d.xlsx' %(filexlsx1, cur.hour, cur.minute)
wb.save(filexlsx1)


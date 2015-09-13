# -*- coding:gbk -*-
import sys
import re
import os
import datetime
import urllib
import urllib2
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

#url = "http://vip.stock.finance.sina.com.cn/quotes_service/view/vMS_tradedetail.php?symbol=sz300001&date=2015-09-10&page=48"
#url = "http://vip.stock.finance.sina.com.cn/quotes_service/view/vMS_tradehistory.php?symbol=sz300001&date=2015-09-10&page=1"
#成交时间	成交价	涨跌幅	价格变动	成交量(手)		成交额(元)	性质
#	<tr ><th>11:29:48</th><td>14.57</td><td>-3.06%</td><td>+0.01</td><td>16</td><td>23,312</td><th><h1>中性盘</h1></th></tr>
#<tr ><th>11:29:36</th><td>14.56</td><td>-3.13%</td><td>--</td><td>9</td><td>13,104</td><th><h6>卖盘</h6></th></tr>
#<tr ><th>11:29:21</th><td>14.56</td><td>-3.13%</td><td>-0.02</td><td>10</td><td>14,560</td><th><h6>卖盘</h6></th></tr>

addcsv = 0
pindex = len(sys.argv)
if (pindex != 3):
	sys.stderr.write("Usage: command 代码 时间<YYYY-MM-DD or MM-DD>\n")
	exit(1);

code = sys.argv[1]
qdate = sys.argv[2]
if (len(code) != 6):
	sys.stderr.write("Len should be 6\n")
	exit(1);

head3 = code[0:3]
result = (cmp(head3, "000")==0) or (cmp(head3, "002")==0) or (cmp(head3, "300")==0)
if result is True:
	code = "sz" + code
else:
	result = (cmp(head3, "600")==0) or (cmp(head3, "601")==0) or (cmp(head3, "603")==0)
	if result is True:
		code = "sh" + code
	else:
		print "非法代码:" +code+ "\n"
		exit(1);
#print code

dateObj = re.match(r'^(\d{4})-(\d+)-(\d+)', qdate)
if (dateObj is None):
	dateObj = re.match(r'^(\d+)-(\d+)', qdate)
	if (dateObj is None):
		print "非法日期格式：" +qdate+ ",期望格式:YYYY-MM-DD or MM-DD"
		exit(1);
	else:
		today = datetime.date.today()
		year = str(today.year)
		month = dateObj.group(1)
		day = dateObj.group(2)
		pass
else:
	year = dateObj.group(1)
	month = dateObj.group(2)
	day = dateObj.group(3)

qdate = year
if len(month)==1:
	qdate = year+ "-0" +month
else:
	qdate = year+ "-" +month
if len(day)==1:
	qdate = qdate+ "-0" +day
else:
	qdate = qdate+ "-" +day
#print qdate

url = "http://vip.stock.finance.sina.com.cn/quotes_service/view/vMS_tradehistory.php?symbol=" + code
url = url + "&date=" + qdate

wb = Workbook()
# grab the active worksheet
ws = wb.active

totalline = 0
lasttime = ''
filename = code+ '_' + qdate
if addcsv==1:
	filecsv = filename + '.csv'
	fcsv = open(filecsv, 'w')
	strline = '成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质'
	fcsv.write(strline)
	fcsv.write("\n")

sellall = 0
buyall = 0
sellct = 0
buyct = 0
sell3all = 0
buy3all = 0
sell3ct = 0
buy3ct = 0
sell6all = 0
buy6all = 0
sell6ct = 0
buy6ct = 0

strline = u'成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质'
strObj = strline.split(u',')
ws.append(strObj)
for i in range(1,1000):
	urlall = url + "&page=" +str(i)
#	print "%d, %s" %(i,urlall)
	
	req = urllib2.Request(urlall)
	res_data = urllib2.urlopen(req)

	flag = 0
	count = 0
	line = res_data.readline()
	checkStr = '成交时间'
	while line:
#		print line
		index = line.find(checkStr)
		if (index<0):
			line = res_data.readline()
			continue

		if flag==0:
			checkStr = '<th>'
			flag = 1
		else:
			key = re.match(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D', line)
			if (key):
#				print key.groups()
				curtime = key.group(1)
				if re.search(curtime, lasttime):
					pass
				else:
					lasttime = curtime
					amount = key.group(6)
					obj = amount.split(',')
					amount = ''.join(obj)
					
					intamount = int(key.group(5))
					state = key.group(7)
					if cmp(state, '卖盘')==0:
						sellall += intamount
						sellct += 1
						if intamount>=300:
							sell3all += intamount
							sell3ct += 1
						if intamount>=600:
							sell6all += intamount
							sell6ct += 1
#						print "S:%d %d" %(sellall, intamount)
					elif cmp(state, '买盘')==0:
						buyall += intamount
						buyct += 1
						if intamount>=300:
							buy3all += intamount
							buy3ct += 1
						if intamount>=600:
							buy6all += intamount
							buy6ct += 1
#						print "B:%d %d" %(buyall, intamount)
						pass
					if addcsv==1:
						strline = curtime +","+ key.group(2) +","+ key.group(3) +","+ key.group(4) +","+ key.group(5) +","+ amount +","+ key.group(7) + "\n"
						fcsv.write(strline)
					
					totalline += 1
					row = totalline+1
					cell = 'A' + str(row)
					ws[cell] = curtime
					cell = 'B' + str(row)
					ws[cell] = key.group(2)
					cell = 'C' + str(row)
					ws[cell] = key.group(3)
					cell = 'D' + str(row)
					ws[cell] = key.group(4)
					cell = 'E' + str(row)
					ws[cell] = int(key.group(5))
					cell = 'F' + str(row)
					ws[cell] = int(amount)
					cell = 'G' + str(row)
					s1 = key.group(7).decode('gbk')
					ws[cell] = s1
				count += 1
				pass
			else:
				endObj = re.search(r'</td><td>', qdate)
				if (endObj):
					print "Error line:" + line
				else:
					break;
		line = res_data.readline()

	if (count==0):
		break;

if addcsv==1:
	fcsv.close()
	if (totalline==0):
		os.remove(filecsv)

ws = wb.create_sheet()
ws.title = 'statistics'

row = 1
cell = 'B' + str(row)
ws[cell] = 'B'
cell = 'C' + str(row)
ws[cell] = 'Bc'
cell = 'D' + str(row)
ws[cell] = 'S'
cell = 'E' + str(row)
ws[cell] = 'Sc'

row = row+1
cell = 'A' + str(row)
ws[cell] = 0
cell = 'B' + str(row)
ws[cell] = buyall
cell = 'C' + str(row)
ws[cell] = buyct
cell = 'D' + str(row)
ws[cell] = sellall
cell = 'E' + str(row)
ws[cell] = sellct

row = row+1
cell = 'A' + str(row)
ws[cell] = 300
cell = 'B' + str(row)
ws[cell] = buy3all
cell = 'C' + str(row)
ws[cell] = buy3ct
cell = 'D' + str(row)
ws[cell] = sell3all
cell = 'E' + str(row)
ws[cell] = sell3ct

row = row+1
cell = 'A' + str(row)
ws[cell] = 600
cell = 'B' + str(row)
ws[cell] = buy6all
cell = 'C' + str(row)
ws[cell] = buy6ct
cell = 'D' + str(row)
ws[cell] = sell6all
cell = 'E' + str(row)
ws[cell] = sell6ct

filexlsx = filename+ '.xlsx'
wb.save(filexlsx)
if (totalline==0):
	print "No Matched Record"
	os.remove(filexlsx)


